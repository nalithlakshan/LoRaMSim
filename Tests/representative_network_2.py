import sys
import argparse
import matplotlib.pyplot as plt
sys.path.append('../')
import loraMeshSimulator as sim
import threading
from openpyxl import load_workbook


# Create a lock for thread-safe operations
excel_lock = threading.Lock()

def append_values_to_excel(file_path, sheet_name, values):
    with excel_lock:
        # Load the workbook and select the sheet
        workbook = load_workbook(file_path)
        
        if sheet_name not in workbook.sheetnames:
            # If the sheet does not exist, create it
            workbook.create_sheet(sheet_name)
        
        sheet = workbook[sheet_name]

        # Append values to the end of the sheet
        sheet.append(values)

        # Save the workbook
        workbook.save(file_path)

def main(repeater_delay_multiplier, avg_send_time, total_sim_packets):
    #-------------------------------------------------------------------------------------
    # Simulation config
    #-------------------------------------------------------------------------------------
    node = sim.node
    nodes = sim.nodes
    env =sim.env
    maxDist =sim.maxDist


    sim.avgSendTime = avg_send_time
    sim.repeatDelayMultiplier = repeater_delay_multiplier
    sim.graphics = 1
    sim.realtime_graphics = 0
    sim.debug = 0

    sim.positional_algo = True
    sim.standby_repeater_algo = True
    sim.energy_aware_algo = True

    sim.totalSimPackets = total_sim_packets

    repeaters =[]
    enddevices = []

    #Repeaters
    d = maxDist*0.99
    
    gw1 = sim.node(env, 6.00*d, 1.00*d, "gw")
    repeaters.append(node(env, 5.50*d, 1.00*d, "rp"))
    repeaters.append(node(env, 5.00*d, 1.00*d, "rp"))
    repeaters.append(node(env, 4.50*d, 1.00*d, "rp"))
    repeaters.append(node(env, 4.00*d, 1.00*d, "rp")) #junction 1
    repeaters.append(node(env, 3.50*d, 1.00*d, "rp")) 
    repeaters.append(node(env, 3.00*d, 1.00*d, "rp"))
    repeaters.append(node(env, 2.66*d, 1.00*d, "rp"))
    repeaters.append(node(env, 2.33*d, 1.00*d, "rp")) #junction 2
    repeaters.append(node(env, 2.00*d, 1.00*d, "rp")) 
    repeaters.append(node(env, 1.66*d, 1.00*d, "rp"))

   
    repeaters.append(node(env, 4.00*d, 1.50*d, "rp")) #branch at j1
    repeaters.append(node(env, 4.00*d, 2.00*d, "rp"))
    repeaters.append(node(env, 4.00*d, 2.50*d, "rp"))

    repeaters.append(node(env, 2.33*d, 1.50*d, "rp")) #branch at j2
    repeaters.append(node(env, 2.33*d, 2.00*d, "rp"))
    repeaters.append(node(env, 2.33*d, 2.33*d, "rp"))
    repeaters.append(node(env, 2.33*d, 2.66*d, "rp"))
    gw2 = sim.node(env, 2.33*d, 3.00*d, "gw")
    
    no_of_repeaters = len(repeaters)


    #End devices
    enddevices.append(node(env, 5.40*d, 0.80*d, "ed"))
    enddevices.append(node(env, 4.90*d, 1.15*d, "ed"))
    enddevices.append(node(env, 4.55*d, 1.20*d, "ed"))
    enddevices.append(node(env, 4.20*d, 0.90*d, "ed")) #junction 1
    enddevices.append(node(env, 3.30*d, 0.90*d, "ed")) 
    enddevices.append(node(env, 3.00*d, 1.20*d, "ed"))
    enddevices.append(node(env, 2.60*d, 1.10*d, "ed"))
    enddevices.append(node(env, 2.25*d, 0.80*d, "ed")) #junction 2
    enddevices.append(node(env, 1.90*d, 1.10*d, "ed")) 
    enddevices.append(node(env, 1.50*d, 0.85*d, "ed"))
   
    enddevices.append(node(env, 4.10*d, 1.40*d, "ed")) #branch at j1
    enddevices.append(node(env, 4.20*d, 2.15*d, "ed"))
    enddevices.append(node(env, 3.80*d, 2.60*d, "ed"))

    enddevices.append(node(env, 2.35*d, 1.60*d, "ed")) #branch at j2
    enddevices.append(node(env, 2.25*d, 1.85*d, "ed"))
    enddevices.append(node(env, 2.55*d, 2.25*d, "ed"))
    enddevices.append(node(env, 2.10*d, 2.60*d, "ed"))

    no_of_enddevices = len(enddevices)


    sim.networkConfig()

    for i in range(19):
        print(f"Node {i}\t", "distanceValue:", round(nodes[i].distanceValue,2),"\tnextRp:",  nodes[i].nextRp, "\tnearest GW:", nodes[i].nearestGwId, "\tsyncLevel:", nodes[i].syncLevel)


    #Sensor Network
    for i in range(len(nodes)):
        if (nodes[i].type.lower() == "ed"):
            env.process(nodes[i].endDeviceStateMachine(env))
        if (nodes[i].type.lower() != "ed"):
            env.process(nodes[i].enableCad(env))

    #prepare show
    if (sim.graphics == 1):
        sim.plt.xlim([0, sim.xmax])
        sim.plt.ylim([0, sim.ymax])
        sim.plt.draw()
        sim.plt.show()

    # start simulation
    env.run()

    #-----------------------------------------------------------------------
    #Print simulation stat
    print ("No of nodes: ", len(nodes)) #FIX
    print ("AvgSendTime (exp. distributed):",sim.avgSendTime)
    print ("Simulation Time: ",env.now/60000,"mins")
    print ("Full Collision: ", sim.full_collision)
    print ("Air time: ", sim.nodes[1].packet[0].rectime)


    #Average Latency
    sum_of_latencies = 0
    sim.packetLatencies.sort()
    for i in range(len(sim.packetLatencies)):
        sum_of_latencies += sim.packetLatencies[i]
    average_latency = sum_of_latencies/len(sim.packetLatencies)

    print("Average Latency:",average_latency,"ms")
    print("Minimum Latency:",sim.packetLatencies[0],"ms")
    print("Maximum Latency:",sim.packetLatencies[-1],"ms")

    print("\n Initial Packet Transmission Successes & Failures by Each End-Device")
    total_pkts_rec_by_bs = 0
    total_ed_tx_pkts = 0
    total_ed_tx_successes = 0
    total_ed_tx_losses = 0
    total_intermediate_losses = 0
    total_power_consumption = 0

    for i in range(no_of_enddevices):
        ed = enddevices[i]
        ed_tx_pkts = 0
        ed_tx_successes = 0

        for pkt in ed.txPackets:
            edId,packetSeq,genTime,pktType =pkt.split("|")
            if(pktType=="DATA_up"):
                total_ed_tx_pkts += 1
                ed_tx_pkts += 1

                ed_tx_successful = 0
                for j in range(no_of_repeaters):
                    rp = repeaters[j]
                    if(pkt in rp.txPackets or pkt in sim.packetsRecBS):
                        ed_tx_successful = 1
                        ed_tx_successes += 1
                        total_ed_tx_successes += 1
                        break

                if(pkt in sim.packetsRecBS):
                    total_pkts_rec_by_bs += 1
                elif(ed_tx_successful):
                    total_intermediate_losses += 1
        
        ed_tx_losses = ed_tx_pkts - ed_tx_successes
        if(ed_tx_pkts>0):
            print("ED", ed.id, "Sent Pkts:", ed_tx_pkts)
            print("ED", ed.id, "Pkts successfully captured by nearby repeater(s):", ed_tx_successes)
            print("ED", ed.id, "Pkts failed to be captured by nearby repeater(s):", ed_tx_losses)
            print("ED", ed.id, "percentage of initial Pkt transmission failures:", round(ed_tx_losses/ed_tx_pkts*100,1), "%\n")
                    
    total_ed_tx_losses = total_ed_tx_pkts - total_ed_tx_successes

    total_tx_power_consumption = 0
    total_rx_power_consumption = 0
    total_cad_power_consumption = 0
    for i in range(no_of_repeaters):
        rp = repeaters[i]
        pkts_rec_by_rp = 0
        pkts_fwd_by_rp = 0
        total_power_consumption += rp.batteryCapacity - rp.batteryRemaining
        total_tx_power_consumption += rp.txPowerConsumption
        total_rx_power_consumption += rp.rxPowerConsumption
        total_cad_power_consumption += rp.cadPowerConsumption



        for pkt in rp.recPackets:
            edId,packetSeq,genTime,pktType =pkt.split("|")
            if(pktType=="DATA_up"):
                pkts_rec_by_rp += 1
        
        for pkt in rp.txPackets:
            edId,packetSeq,genTime,pktType =pkt.split("|")
            if(pktType=="DATA_up"):
                pkts_fwd_by_rp += 1

        if(len(rp.recPackets) <10):
            print("RP", rp.id, "Received Pkts:", rp.recPackets)
        print("RP", rp.id, "Received Pkts:", pkts_rec_by_rp)
        print("RP", rp.id, "Repeated Pkts:", pkts_fwd_by_rp)
        print("RP", rp.id, "Battery Consumed:", round(rp.batteryCapacity - rp.batteryRemaining,2), "mAh\n")
    
    print("********************************************")
    print("Total Generated Packets:", total_sim_packets)

    total_lost_pkts = total_sim_packets - total_pkts_rec_by_bs

    print("Total Lost Packets:", total_lost_pkts)
    print("---> Lost at initial ED transmission:", total_ed_tx_losses)
    print("---> Lost at intermediate repetition:", total_intermediate_losses, "\n")

    print("\nTotal Standy \t:",sim.total_stanby)
    print("---> Standby Retains \t\t:",sim.standby_retains)
    print("---> Standby Recoveries \t:",sim.standby_recoveries,"\n")
    print("---> Energy Aware Repeater Role Changes \t:",sim.repeater_role_changes)

    print("\nTotal Power Consumption \t:", total_power_consumption,"mAh")
    print("---> Total Tx Power Consumption :", total_tx_power_consumption, "mAh")
    print("---> Total Rx Power Consumption :", total_rx_power_consumption, "mAh")
    print("---> Total CAD Power Consumption:", total_cad_power_consumption, "mAh")


    # data extraction rate
    der = total_pkts_rec_by_bs/float(total_sim_packets)
    print("\nOverall DER:", der)



    # Append Test to Excel Sheet
    file_path = "representative_network_2_sim_outputs.xlsx"
    sheet_name = "Sheet1"
    values_to_append = []
    values_to_append.append(sim.experiment)
    values_to_append.append(no_of_repeaters)
    values_to_append.append(repeater_delay_multiplier)
    values_to_append.append(avg_send_time)
    values_to_append.append(total_sim_packets)
    values_to_append.append("---> DER:")
    values_to_append.append(der)
    values_to_append.append("---> Avg Latency:")
    values_to_append.append(average_latency)
    values_to_append.append("---> Min Latency:")
    values_to_append.append(sim.packetLatencies[0])
    values_to_append.append("---> Max Latency:")
    values_to_append.append(sim.packetLatencies[-1])
    values_to_append.append("---> Total Power:")
    values_to_append.append(total_power_consumption)
    values_to_append.append(" TX:")
    values_to_append.append(total_tx_power_consumption)
    values_to_append.append(" RX:")
    values_to_append.append(total_rx_power_consumption)
    values_to_append.append(" CAD:")
    values_to_append.append(total_cad_power_consumption)
    # values_to_append.append("---> Total Collisions:")
    # values_to_append.append(len(sim.collidedPackets))
    # values_to_append.append("---> Total Lost Pkts:")
    # values_to_append.append(total_lost_pkts)
    # values_to_append.append("---> Lost at initial ED transmission:")
    # values_to_append.append(total_ed_tx_losses)
    # values_to_append.append("---> Lost at intermediate repetition:")
    # values_to_append.append(total_intermediate_losses)

    append_values_to_excel(file_path, sheet_name, values_to_append)

    # this can be done to keep graphics visible
    if (sim.graphics == 1):
        input('Press Enter to continue ...')


#============================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Repeater Waiting Period Simulation")
    
    # # Add your command-line arguments
    parser.add_argument("-repeater_delay_multiplier" , default=3 , help="How many times the repeater mean waiting period is greater than the pkt transmission air time?")
    parser.add_argument("-avg_send_time"             , default=36000000 , help="Average time period of an end-device sending a packet")
    parser.add_argument("-total_sim_packets"         , default=1000 , help="Total number of packets to process in the simulation")

    # # Parse the command-line arguments
    args = parser.parse_args()

    # Call the main function with the parsed arguments
    main(int(args.repeater_delay_multiplier), int(args.avg_send_time), int(args.total_sim_packets))
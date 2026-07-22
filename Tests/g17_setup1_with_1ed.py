import sys
import argparse
import os
import matplotlib.pyplot as plt
sys.path.append('../')
import loraMeshSimulator as sim
import threading
from openpyxl import Workbook, load_workbook


# Create a lock for thread-safe operations
excel_lock = threading.Lock()

def append_values_to_excel(file_path, sheet_name, values):
    with excel_lock:
        # Load the workbook and select the sheet
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
        else:
            workbook = Workbook()
            workbook.active.title = sheet_name
        
        if sheet_name not in workbook.sheetnames:
            # If the sheet does not exist, create it
            workbook.create_sheet(sheet_name)
        
        sheet = workbook[sheet_name]

        # Append values to the end of the sheet
        sheet.append(values)

        # Save the workbook
        workbook.save(file_path)

def main(repeater_delay_multiplier, avg_send_time, total_sim_packets):
    #-------------------------------------------------------------------------------------
    # Simulation config
    #-------------------------------------------------------------------------------------
    node = sim.node
    nodes = sim.nodes
    env =sim.env
    maxDist =sim.maxDist


    sim.avgSendTime = avg_send_time
    sim.repeatDelayMultiplier = repeater_delay_multiplier
    sim.graphics = 1
    sim.realtime_graphics = 1
    sim.debug = 0

    sim.positional_algo = True
    sim.standby_repeater_algo = True
    sim.energy_aware_algo = False

    sim.totalSimPackets = total_sim_packets

    repeaters =[]
    enddevices = []

    # The graph uses one-based labels, while LoRaMSim assigns zero-based IDs.
    # Infrastructure nodes must be created before EDs for position learning.
    d = maxDist*0.99

    gw1 = node(env, 2.90*d, 1.40*d, "gw")             # graph ID 1
    rp3 = node(env, 2.00*d, 1.40*d, "rp")             # graph ID 3
    rp4 = node(env, 1.00*d, 1.40*d, "rp")             # graph ID 4
    rp5 = node(env, 1.00*d, 1.00*d, "rp")             # graph ID 5
    rp6 = node(env, 2.00*d, 1.00*d, "rp")             # graph ID 6
    rp7 = node(env, 3.00*d, 1.00*d, "rp")             # graph ID 7
    repeaters.extend([rp3, rp4, rp5, rp6, rp7])

    no_of_repeaters = len(repeaters)

    ed22 = node(env, 0.90*d, 1.50*d, "ed")             # graph ID 22, near RP 4
    enddevices.append(ed22)


    topology_nodes = {
        1: gw1,
        22: ed22,
        3: rp3,
        4: rp4,
        5: rp5,
        6: rp6,
        7: rp7,
    }
    topology_id_by_sim_id = {
        topology_node.id: topology_id
        for topology_id, topology_node in topology_nodes.items()
    }

    if sim.graphics == 1:
        for topology_id, topology_node in topology_nodes.items():
            topology_node.label.set_text(topology_id)
    

    sim.networkConfig()

    print("\nTopology ID -> simulator ID")
    for topology_id in sorted(topology_nodes):
        print(topology_id, "->", topology_nodes[topology_id].id)

    print("\nPosition-learning results")
    for topology_id in [1, 3, 4, 5, 6, 7]:
        topology_node = topology_nodes[topology_id]
        next_rp = topology_id_by_sim_id.get(topology_node.nextRp, topology_node.nextRp)
        nearest_gw = topology_id_by_sim_id.get(
            topology_node.nearestGwId, topology_node.nearestGwId
        )
        print(
            "Node", topology_id,
            "distanceValue:", topology_node.distanceValue,
            "nextRp:", next_rp,
            "nearest GW:", nearest_gw,
        )


    #Sensor Network
    for i in range(len(nodes)):
        if (nodes[i].type.lower() == "ed"):
            env.process(nodes[i].endDeviceStateMachine(env))
        if (nodes[i].type.lower() == "rp"):
            env.process(nodes[i].enableCad(env))


    #prepare show
    if (sim.graphics == 1):
        sim.plt.xlim([0, sim.xmax])
        sim.plt.ylim([0, sim.ymax])
        sim.plt.draw()
        sim.plt.show()

    # start simulation
    env.run()

    #-----------------------------------------------------------------------
    #Print simulation stat
    print ("No of nodes: ", len(nodes)) #FIX
    print ("AvgSendTime (exp. distributed):",sim.avgSendTime)
    print ("Simulation Time: ",env.now/60000,"mins")
    print ("Full Collision: ", sim.full_collision)
    print ("Air time: ", gw1.packet[0].rectime)

    # data extraction rate
    der = len(sim.packetsRecBS)/float(sim.totalSimPackets)
    print("\nOverall DER:", der)

    # Average latency. A zero-delivery run has no latency samples.
    sim.packetLatencies.sort()
    if sim.packetLatencies:
        average_latency = sum(sim.packetLatencies)/len(sim.packetLatencies)
        minimum_latency = sim.packetLatencies[0]
        maximum_latency = sim.packetLatencies[-1]
        print("Average Latency:", average_latency, "ms")
        print("Minimum Latency:", minimum_latency, "ms")
        print("Maximum Latency:", maximum_latency, "ms")
    else:
        average_latency = None
        minimum_latency = None
        maximum_latency = None
        print("Average Latency: N/A (no packets reached the gateway)")
        print("Minimum Latency: N/A")
        print("Maximum Latency: N/A")

    print("\n Received/Repeated Packets by Each Repeater")
    total_ed_tx_successes = 0
    total_ed_tx_losses = 0
    total_power_consumption = 0
    # This topology has one ED and it is near repeater 4.
    ed_tx_pkts = len(ed22.txPackets)
    ed_tx_successes = len([x for x in ed22.txPackets if x in rp4.recPackets])
    ed_tx_losses = ed_tx_pkts - ed_tx_successes
    total_ed_tx_successes += ed_tx_successes
    total_ed_tx_losses += ed_tx_losses
    if ed_tx_pkts > 0:
        if ed_tx_pkts < 5:
            print("ED", 22, "Sent Pkts:", ed22.txPackets)
        print("ED", 22, "Sent Pkts:", ed_tx_pkts)
        print("ED", 22, "Pkts successfully sent to corresponding repeater:", ed_tx_successes)
        print("ED", 22, "Pkts failed to be captured by corresponding repeater:", ed_tx_losses)
        print("ED", 22, "percentage of initial Pkt transmission failures:", round(ed_tx_losses/ed_tx_pkts*100,1), "%")

    for topology_id, rp in zip([3, 4, 5, 6, 7], repeaters):
        total_power_consumption += rp.batteryCapacity - rp.batteryRemaining
        if(len(rp.recPackets) <20):
            print("RP", topology_id, "Received Pkts:", rp.recPackets)
        print("RP", topology_id, "Received Pkts:", len(rp.recPackets))
        print("RP", topology_id, "Repeated Pkts:", len(rp.txPackets))
        # print("RP", rp.id, "percentage time RP was in Tx state:", round(rp.packet[0].rectime*len(rp.txPackets)/env.now*100,1),"%\n")
        

    
    print("********************************************")
    print("Total Generated Packets:", total_sim_packets)
    # total_lost_pkts = total_sim_packets - len(set(gw1.recPackets+gw2.recPackets))
    total_lost_pkts = total_sim_packets - len(sim.packetsRecBS)

    total_intermediate_losses = total_lost_pkts-total_ed_tx_losses
    print("Total Lost Packets:", total_lost_pkts)
    print("---> Lost at initial ED transmission:", total_ed_tx_losses)
    print("---> Lost at intermediate repetition:", total_intermediate_losses, "\n")

    print("\nTotal Standy \t:",sim.total_stanby)
    print("---> Standby Retains \t\t:",sim.standby_retains)
    print("---> Standby Recoveries \t:",sim.standby_recoveries,"\n")
    print("---> Energy Aware Repeater Role Changes \t:",sim.repeater_role_changes,"\n")

    print("\nTotal Power Consumption :", total_power_consumption)

    print("\nOverall DER:", der)

    #print RSSI of packets
    pkt1_tx = 0
    pkt1_rx = 1

    pkt2_tx = 3
    pkt2_rx = 1
    # print("Node",pkt1_tx,"to Node",pkt1_rx,"RSSI:", nodes[pkt1_tx].packet[pkt1_rx].rssi)
    # print("Node",pkt2_tx,"to Node",pkt2_rx,"RSSI:", nodes[pkt2_tx].packet[pkt2_rx].rssi)

    # Append Test to Excel Sheet
    file_path = "g17_setup1_with_1ed_sim_outputs.xlsx"
    sheet_name = "Sheet1"
    values_to_append = []
    values_to_append.append(sim.experiment)
    values_to_append.append(no_of_repeaters)
    values_to_append.append(repeater_delay_multiplier)
    values_to_append.append(avg_send_time)
    values_to_append.append(total_sim_packets)
    values_to_append.append("---> DER:")
    values_to_append.append(der)
    values_to_append.append("---> Avg Latency:")
    values_to_append.append(average_latency)
    values_to_append.append("---> Min Latency:")
    values_to_append.append(minimum_latency)
    values_to_append.append("---> Max Latency:")
    values_to_append.append(maximum_latency)
    # values_to_append.append("---> Total Collisions:")
    # values_to_append.append(len(sim.collidedPackets))
    # values_to_append.append("---> Total Lost Pkts:")
    # values_to_append.append(total_lost_pkts)
    # values_to_append.append("---> Lost at initial ED transmission:")
    # values_to_append.append(total_ed_tx_losses)
    # values_to_append.append("---> Lost at intermediate repetition:")
    # values_to_append.append(total_intermediate_losses)

    append_values_to_excel(file_path, sheet_name, values_to_append)

    # this can be done to keep graphics visible
    if (sim.graphics == 1):
        input('Press Enter to continue ...')


#============================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Repeater Waiting Period Simulation")
    
    # # Add your command-line arguments
    parser.add_argument("-repeater_delay_multiplier" , default=3 , help="How many times the repeater mean waiting period is greater than the pkt transmission air time?")
    parser.add_argument("-avg_send_time"             , default=3600000 , help="Average time period of an end-device sending a packet")
    parser.add_argument("-total_sim_packets"         , default=3 , help="Total number of packets to process in the simulation")

    # # Parse the command-line arguments
    args = parser.parse_args()

    # Call the main function with the parsed arguments
    main(int(args.repeater_delay_multiplier), int(args.avg_send_time), int(args.total_sim_packets))

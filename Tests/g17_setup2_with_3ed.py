import sys
import argparse
import os
import matplotlib.pyplot as plt
sys.path.append('../')
import loraMeshSimulator as sim
import threading
from openpyxl import Workbook, load_workbook


# Create a lock for thread-safe operations
excel_lock = threading.Lock()

def append_values_to_excel(file_path, sheet_name, values):
    with excel_lock:
        # Load the workbook and select the sheet
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
        else:
            workbook = Workbook()
            workbook.active.title = sheet_name
        
        if sheet_name not in workbook.sheetnames:
            # If the sheet does not exist, create it
            workbook.create_sheet(sheet_name)
        
        sheet = workbook[sheet_name]

        # Append values to the end of the sheet
        sheet.append(values)

        # Save the workbook
        workbook.save(file_path)

def main(repeater_delay_multiplier, avg_send_time, total_sim_packets):
    #-------------------------------------------------------------------------------------
    # Simulation config
    #-------------------------------------------------------------------------------------
    node = sim.node
    nodes = sim.nodes
    env =sim.env
    maxDist =sim.maxDist


    sim.avgSendTime = avg_send_time
    sim.repeatDelayMultiplier = repeater_delay_multiplier
    sim.graphics = 0
    sim.realtime_graphics = 0
    sim.debug = 0

    sim.positional_algo = True
    sim.standby_repeater_algo = True
    sim.energy_aware_algo = False

    sim.totalSimPackets = total_sim_packets

    repeaters =[]
    enddevices = []

    # The graph uses one-based labels, while LoRaMSim assigns zero-based IDs.
    # Infrastructure nodes must be created before EDs for position learning.
    d = maxDist*0.99

    gw1 = node(env, 2.90*d, 1.80*d, "gw")             # graph ID 1
    rp3 = node(env, 2.00*d, 1.50*d, "rp")             # graph ID 2
    rp4 = node(env, 1.00*d, 1.50*d, "rp")             # graph ID 3
    rp5 = node(env, 1.00*d, 1.00*d, "rp")             # graph ID 4
    rp6 = node(env, 2.00*d, 1.00*d, "rp")             # graph ID 5
    rp7 = node(env, 2.90*d, 0.70*d, "rp")             # graph ID 6
    rp8 = node(env, 2.50*d, 1.80*d, "rp")             # graph ID 7
    rp10 = node(env, 2.50*d, 0.70*d, "rp")            # graph ID 10
    rp9 = node(env, 1.50*d, 1.80*d, "rp")             # graph ID 8
    rp2 = node(env, 1.50*d, 0.70*d, "rp")             # graph ID 9
    repeaters.extend([rp3, rp4, rp5, rp6, rp7, rp8, rp10, rp9, rp2])
    repeater_topology_ids = [2, 3, 4, 5, 6, 7, 10, 8, 9]

    no_of_repeaters = len(repeaters)

    ed22 = node(env, 0.90*d, 1.60*d, "ed")             # graph ID 22, near RP 3
    ed23 = node(env, 2.00*d, 1.65*d, "ed")             # graph ID 23, near RP 2
    ed24 = node(env, 2.90*d, 0.55*d, "ed")             # graph ID 24, near RP 6
    
    enddevices.extend([ed22, ed23, ed24])


    topology_nodes = {
        1: gw1,
        22: ed22,
        23: ed23,
        24: ed24,
        2: rp3,
        3: rp4,
        4: rp5,
        5: rp6,
        6: rp7,
        7: rp8,
        10: rp10,
        8: rp9,
        9: rp2,
    }
    topology_id_by_sim_id = {
        topology_node.id: topology_id
        for topology_id, topology_node in topology_nodes.items()
    }

    if sim.graphics == 1:
        device_type_prefix = {
            "gw": "G",
            "rp": "R",
            "ed": "E",
        }
        for topology_id, topology_node in topology_nodes.items():
            prefix = device_type_prefix[topology_node.type.lower()]
            topology_node.label.set_text(
                f"{topology_node.id} ({prefix}{topology_id})"
            )
    

    sim.networkConfig()

    print("\nTopology ID -> simulator ID")
    for topology_id in sorted(topology_nodes):
        print(topology_id, "->", topology_nodes[topology_id].id)

    print("\nPosition-learning results")
    for topology_id in [1] + sorted(repeater_topology_ids):
        topology_node = topology_nodes[topology_id]
        next_rp = topology_id_by_sim_id.get(topology_node.nextRp, topology_node.nextRp)
        nearest_gw = topology_id_by_sim_id.get(
            topology_node.nearestGwId, topology_node.nearestGwId
        )
        print(
            "Node", topology_id,
            "distanceValue:", topology_node.distanceValue,
            "nextRp:", next_rp,
            "nearest GW:", nearest_gw,
        )


    #Sensor Network
    for i in range(len(nodes)):
        if (nodes[i].type.lower() == "ed"):
            env.process(nodes[i].endDeviceStateMachine(env))
        if (nodes[i].type.lower() == "rp"):
            env.process(nodes[i].enableCad(env))


    #prepare show
    if (sim.graphics == 1):
        sim.plt.xlim([0, sim.xmax])
        sim.plt.ylim([0, sim.ymax])
        sim.plt.draw()
        sim.plt.show()

    # start simulation
    env.run()

    #-----------------------------------------------------------------------
    #Print simulation stat
    print ("No of nodes: ", len(nodes)) #FIX
    print ("AvgSendTime (exp. distributed):",sim.avgSendTime)
    print ("Simulation Time: ",env.now/60000,"mins")
    print ("Full Collision: ", sim.full_collision)
    print ("Air time: ", gw1.packet[0].rectime)

    # data extraction rate
    der = len(sim.packetsRecBS)/float(sim.totalSimPackets)
    print("\nOverall DER:", der)

    # Average latency. A zero-delivery run has no latency samples.
    sim.packetLatencies.sort()
    if sim.packetLatencies:
        average_latency = sum(sim.packetLatencies)/len(sim.packetLatencies)
        minimum_latency = sim.packetLatencies[0]
        maximum_latency = sim.packetLatencies[-1]
        print("Average Latency:", average_latency, "ms")
        print("Minimum Latency:", minimum_latency, "ms")
        print("Maximum Latency:", maximum_latency, "ms")
    else:
        average_latency = None
        minimum_latency = None
        maximum_latency = None
        print("Average Latency: N/A (no packets reached the gateway)")
        print("Minimum Latency: N/A")
        print("Maximum Latency: N/A")

    print("\nReceived/Repeated Packets by Each Repeater")
    total_ed_tx_successes = 0
    total_ed_tx_losses = 0
    total_power_consumption = 0
    generated_data_packets = set()
    initial_data_losses_by_corresponding_rp = set()

    # ED 22 is near RP 3, ED 23 is near RP 2, and ED 24 is near RP 6.
    ed_repeater_pairs = [
        (22, ed22, rp4),
        (23, ed23, rp3),
        (24, ed24, rp7),
    ]
    for topology_id, ed, corresponding_rp in ed_repeater_pairs:
        ed_wor_packets = [
            packet for packet in ed.txPackets
            if packet.endswith("|WOR_up")
        ]
        ed_data_packets = [
            packet for packet in ed.txPackets
            if packet.endswith("|DATA_up")
        ]
        generated_data_packets.update(ed_data_packets)

        ed_wor_successes = len([
            packet for packet in ed_wor_packets
            if packet in corresponding_rp.recPackets
        ])
        ed_data_successes = len([
            packet for packet in ed_data_packets
            if packet in corresponding_rp.recPackets
        ])
        ed_wor_losses = len(ed_wor_packets) - ed_wor_successes
        ed_data_losses = len(ed_data_packets) - ed_data_successes
        initial_data_losses_by_corresponding_rp.update(
            packet for packet in ed_data_packets
            if packet not in corresponding_rp.recPackets
        )
        total_ed_tx_successes += ed_data_successes
        total_ed_tx_losses += ed_data_losses

        print("ED", topology_id, "WOR sent/received/lost:",
              len(ed_wor_packets), ed_wor_successes, ed_wor_losses)
        print("ED", topology_id, "DATA sent/received/lost:",
              len(ed_data_packets), ed_data_successes, ed_data_losses)
        if ed_data_packets:
            print(
                "ED", topology_id,
                "percentage of initial DATA transmission failures:",
                round(ed_data_losses/len(ed_data_packets)*100, 3), "%"
            )

    for topology_id, rp in zip(repeater_topology_ids, repeaters):
        total_power_consumption += rp.batteryCapacity - rp.batteryRemaining
        if(len(rp.recPackets) <20):
            print("RP", topology_id, "Received Pkts:", rp.recPackets)
        print("RP", topology_id, "Received Pkts:", len(rp.recPackets))
        print("RP", topology_id, "Repeated Pkts:", len(rp.txPackets))
        # print("RP", rp.id, "percentage time RP was in Tx state:", round(rp.packet[0].rectime*len(rp.txPackets)/env.now*100,1),"%\n")
        

    
    print("********************************************")
    print("Total Generated Packets:", total_sim_packets)
    gateway_data_packets = set(sim.packetsRecBS)
    missing_gateway_packets = sorted(generated_data_packets - gateway_data_packets)
    total_lost_pkts = len(missing_gateway_packets)

    initial_loss_packets = []
    intermediate_loss_packets = []
    for seq_nr in missing_gateway_packets:
        diagnostic = sim.dataPacketDiagnostics.get(seq_nr, {})
        source_id = diagnostic.get("source", int(seq_nr.split("|")[0]))
        source_transmissions = [
            transmission
            for transmission in diagnostic.get("transmissions", [])
            if transmission["sender"] == source_id
        ]
        first_hop_received = any(
            receiver["outcome"] == "scheduled_receive"
            for transmission in source_transmissions
            for receiver in transmission["receivers"]
            if receiver["receiverType"].lower() in ("rp", "gw")
        )
        if first_hop_received:
            intermediate_loss_packets.append(seq_nr)
        else:
            initial_loss_packets.append(seq_nr)

    total_intermediate_losses = len(intermediate_loss_packets)
    print("Total Lost Packets:", total_lost_pkts)
    print("---> Lost at initial ED DATA transmission:", len(initial_loss_packets))
    print("---> Lost at intermediate repetition:", total_intermediate_losses, "\n")

    if missing_gateway_packets:
        print("Missing DATA packet path diagnostics")
        for seq_nr in missing_gateway_packets:
            diagnostic = sim.dataPacketDiagnostics.get(seq_nr, {})
            source_id = diagnostic.get("source", int(seq_nr.split("|")[0]))
            source_label = topology_id_by_sim_id.get(source_id, source_id)
            loss_stage = (
                "initial ED transmission"
                if seq_nr in initial_loss_packets
                else "intermediate repetition"
            )
            print(
                f"\n{seq_nr}: source={source_id} (E{source_label}), "
                f"stage={loss_stage}, "
                f"intendedFirstHop={diagnostic.get('intendedFirstHop')}"
            )
            logical_packet_id = "|".join(seq_nr.split("|")[:2])
            wor_diagnostic = sim.worPacketDiagnostics.get(
                logical_packet_id, {}
            )
            for transmission in wor_diagnostic.get("transmissions", []):
                for receiver in transmission["receivers"]:
                    if receiver["outcome"] == "out_of_range":
                        continue
                    receiver_id = receiver["receiver"]
                    receiver_label = topology_id_by_sim_id.get(
                        receiver_id, receiver_id
                    )
                    print(
                        f"  Initial WOR to node "
                        f"{receiver_id}({receiver_label}): "
                        f"{receiver['outcome']}"
                        f"[startMode={receiver['startMode']},"
                        f"endMode={receiver['endMode']},"
                        f"cadActive={receiver['cadProcessActive']},"
                        f"scanDuringPreamble="
                        f"{receiver['scanOccurredDuringPreamble']},"
                        f"lastScan={receiver['lastCadScanTimeAtEnd']:.2f},"
                        f"preambleEnd={transmission['preambleEnd']:.2f}]"
                    )
            for transmission in diagnostic.get("transmissions", []):
                sender_id = transmission["sender"]
                sender_label = topology_id_by_sim_id.get(sender_id, sender_id)
                visible_results = [
                    receiver
                    for receiver in transmission["receivers"]
                    if receiver["outcome"] != "out_of_range"
                ]
                formatted_results = []
                for receiver in visible_results:
                    receiver_id = receiver["receiver"]
                    receiver_label = topology_id_by_sim_id.get(
                        receiver_id, receiver_id
                    )
                    formatted_results.append(
                        f"{receiver_id}({receiver_label}):"
                        f"{receiver['outcome']}"
                        f"[mode={receiver['mode']},"
                        f"cadActive={receiver['cadProcessActive']},"
                        f"collided={receiver['collided']},"
                        f"processed={receiver['processed']}]"
                    )
                print(
                    f"  TX node {sender_id}({sender_label}) at "
                    f"{transmission['start']:.2f}: "
                    + (", ".join(formatted_results) or "no in-range infrastructure")
                )
            for receive_event in diagnostic.get("receiveEvents", []):
                receiver_id = receive_event["receiver"]
                receiver_label = topology_id_by_sim_id.get(
                    receiver_id, receiver_id
                )
                packet_next_rp = receive_event.get("packetNextRp")
                packet_next_label = topology_id_by_sim_id.get(
                    packet_next_rp, packet_next_rp
                )
                receiver_next_rp = receive_event.get("receiverNextRp")
                receiver_next_label = topology_id_by_sim_id.get(
                    receiver_next_rp, receiver_next_rp
                )
                print(
                    f"  RX decision at node {receiver_id}({receiver_label}) at "
                    f"{receive_event['time']:.2f}: "
                    f"{receive_event.get('decision')}; "
                    f"WOR ACK={receive_event.get('worAckReceived')}, "
                    f"packet next RP={packet_next_rp}({packet_next_label}), "
                    f"receiver next RP={receiver_next_rp}({receiver_next_label}), "
                    f"standby packets={receive_event.get('standbyBufferCount')}"
                )

    print("\nTotal Standy \t:",sim.total_stanby)
    print("---> Standby Retains \t\t:",sim.standby_retains)
    print("---> Standby Recoveries \t:",sim.standby_recoveries,"\n")
    print("---> Energy Aware Repeater Role Changes \t:",sim.repeater_role_changes,"\n")

    print("\nTotal Power Consumption :", total_power_consumption)

    print("\nOverall DER:", der)

    #print RSSI of packets
    pkt1_tx = 0
    pkt1_rx = 1

    pkt2_tx = 3
    pkt2_rx = 1
    # print("Node",pkt1_tx,"to Node",pkt1_rx,"RSSI:", nodes[pkt1_tx].packet[pkt1_rx].rssi)
    # print("Node",pkt2_tx,"to Node",pkt2_rx,"RSSI:", nodes[pkt2_tx].packet[pkt2_rx].rssi)

    # Append Test to Excel Sheet
    file_path = "g17_setup2_with_3ed_sim_outputs.xlsx"
    sheet_name = "Sheet1"
    values_to_append = []
    values_to_append.append(sim.experiment)
    values_to_append.append(no_of_repeaters)
    values_to_append.append(repeater_delay_multiplier)
    values_to_append.append(avg_send_time)
    values_to_append.append(total_sim_packets)
    values_to_append.append("---> DER:")
    values_to_append.append(der)
    values_to_append.append("---> Avg Latency:")
    values_to_append.append(average_latency)
    values_to_append.append("---> Min Latency:")
    values_to_append.append(minimum_latency)
    values_to_append.append("---> Max Latency:")
    values_to_append.append(maximum_latency)
    # values_to_append.append("---> Total Collisions:")
    # values_to_append.append(len(sim.collidedPackets))
    # values_to_append.append("---> Total Lost Pkts:")
    # values_to_append.append(total_lost_pkts)
    # values_to_append.append("---> Lost at initial ED transmission:")
    # values_to_append.append(total_ed_tx_losses)
    # values_to_append.append("---> Lost at intermediate repetition:")
    # values_to_append.append(total_intermediate_losses)

    append_values_to_excel(file_path, sheet_name, values_to_append)

    # this can be done to keep graphics visible
    if (sim.graphics == 1):
        input('Press Enter to continue ...')


#============================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Repeater Waiting Period Simulation")
    
    # # Add your command-line arguments
    parser.add_argument("-repeater_delay_multiplier" , default=3 , help="How many times the repeater mean waiting period is greater than the pkt transmission air time?")
    parser.add_argument("-avg_send_time"             , default=3600000 , help="Average time period of an end-device sending a packet")
    parser.add_argument("-total_sim_packets"         , default=10000 , help="Total number of packets to process in the simulation")

    # # Parse the command-line arguments
    args = parser.parse_args()

    # Call the main function with the parsed arguments
    main(int(args.repeater_delay_multiplier), int(args.avg_send_time), int(args.total_sim_packets))
